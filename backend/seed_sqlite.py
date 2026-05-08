"""
Fleet Command - SQLite Data Seeder
Reads orders_master.xlsx and seeds the local SQLite fleet.db.
Run:  python backend/seed_sqlite.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime
from database import init_db, get_db, log_event

CITY_COORDS = {
    'Chennai':         (13.0827, 80.2707),
    'Kochi':           (9.9312,  76.2673),
    'Mumbai':          (19.0760, 72.8777),
    'Goa':             (15.2993, 74.1240),
    'Kolkata':         (22.5726, 88.3639),
    'Guwahati':        (26.1445, 91.7362),
    'Visakhapatnam':   (17.6868, 83.2185),
    'Bhubaneswar':     (20.2961, 85.8245),
    'Tuticorin':       (8.7642,  78.1348),
    'Trivandrum':      (8.5241,  76.9366),
    'Bengaluru':       (12.9716, 77.5946),
    'Hyderabad':       (17.3850, 78.4867),
}

def find_xlsx():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(project_root, 'orders_master.xlsx'),
        os.path.join(os.getcwd(), 'orders_master.xlsx'),
        os.path.join(project_root, 'data', 'orders_master.xlsx'),
        os.path.join(os.path.dirname(__file__), 'orders_master.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(
        f"\n[ERROR] orders_master.xlsx not found!\n"
        f"  Searched: {candidates}\n"
    )

def load_xlsx(path):
    print(f"   Reading: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    def sheet_rows(name):
        ws = wb[name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for ci, h in enumerate(headers):
                v = ws.cell(r, ci + 1).value
                if hasattr(v, 'isoformat'):
                    v = v.isoformat()
                row[h] = v
            if any(v is not None for v in row.values()):
                rows.append(row)
        return rows

    return {
        'orders':      sheet_rows('orders_master'),
        'tracking':    sheet_rows('order_tracking'),
        'alerts':      sheet_rows('order_alerts'),
        'documents':   sheet_rows('order_documents'),
        'temperature': sheet_rows('temperature_logs'),
    }

def wipe_tables(conn):
    """Delete all rows from core tables (SQLite compatible)."""
    for tbl in ('vehicle_positions', 'temperature_logs', 'order_documents',
                'alerts', 'system_logs', 'orders', 'vehicles'):
        try:
            conn.execute(f"DELETE FROM {tbl}")
        except Exception as e:
            print(f"   [WARN] Could not wipe {tbl}: {e}")
    conn.commit()
    print("   Wiped existing data.")

def seed():
    print(f"\nFleet Command — SQLite Database Seeder")
    print("=" * 45)

    xlsx_path = find_xlsx()

    print("   Initializing database schema...")
    init_db()

    conn = get_db()
    wipe_tables(conn)

    data       = load_xlsx(xlsx_path)
    orders     = data['orders']
    tracking   = {t['Order_ID']: t for t in data['tracking']}
    order_vmap = {o['Order_ID']: o['Vehicle_ID'] for o in orders}

    # ── Build vehicles dict ───────────────────────────────
    vehicles = {}
    for o in orders:
        vid = o['Vehicle_ID']
        if vid and vid not in vehicles:
            vehicles[vid] = {
                'id':               vid,
                'vehicle_type':     o.get('Vehicle_Type', ''),
                'vehicle_number':   o.get('Vehicle_Number', ''),
                'driver_name':      o.get('Driver_Name', ''),
                'driver_contact':   o.get('Driver_Contact', ''),
                'status':           'idle',
                'current_lat':      None,
                'current_lng':      None,
                'current_speed':    0,
                'current_order_id': None,
                'assigned_route':   None,
            }

    # ── Set positions for In Transit vehicles ─────────────
    for o in orders:
        vid = o['Vehicle_ID']
        if not vid or vid not in vehicles:
            continue
        src = o.get('Source_City', '')
        dst = o.get('Destination_City', '')
        t   = tracking.get(o['Order_ID'], {})

        if o.get('Order_Status') == 'In Transit':
            src_c = CITY_COORDS.get(src, (20.5937, 78.9629))
            dst_c = CITY_COORDS.get(dst, (20.5937, 78.9629))
            dist_rem   = float(t.get('Distance_Remaining_km') or 0)
            dist_total = float(o.get('Distance_km') or 1) or 1
            progress   = max(0.0, min(1.0, 1 - dist_rem / dist_total))
            lat = src_c[0] + (dst_c[0] - src_c[0]) * progress
            lng = src_c[1] + (dst_c[1] - src_c[1]) * progress
            vehicles[vid].update({
                'status':           'moving',
                'current_lat':      round(lat, 6),
                'current_lng':      round(lng, 6),
                'current_speed':    float(t.get('Current_Speed_kmph') or 45),
                'current_order_id': o['Order_ID'],
                'assigned_route':   f"{src} → {dst}",
            })
        elif o.get('Order_Status') == 'Delivered' and vehicles[vid]['current_lat'] is None:
            dst_c = CITY_COORDS.get(dst, (13.0827, 80.2707))
            vehicles[vid].update({
                'status':      'idle',
                'current_lat': dst_c[0],
                'current_lng': dst_c[1],
            })

    # ── Insert vehicles (SQLite uses ? placeholders) ──────
    now = datetime.now().isoformat()
    conn.executemany("""
        INSERT OR REPLACE INTO vehicles
            (id, vehicle_type, vehicle_number, plate_number,
             driver_name, driver_contact, status,
             current_lat, current_lng, current_speed,
             current_order_id, assigned_route, last_updated)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [
        (v['id'], v['vehicle_type'], v['vehicle_number'], v['vehicle_number'],
         v['driver_name'], v['driver_contact'], v['status'],
         v['current_lat'], v['current_lng'], v['current_speed'],
         v['current_order_id'], v['assigned_route'], now)
        for v in vehicles.values()
    ])
    print(f"   [OK] Vehicles inserted   : {len(vehicles)}")

    # ── Insert orders ─────────────────────────────────────
    conn.executemany("""
        INSERT OR REPLACE INTO orders
            (id, customer_name, customer_company, customer_contact,
             pickup_address, delivery_address,
             source_city, source_state, source_pincode,
             destination_city, destination_state, destination_pincode,
             goods_type, goods_category, quantity, unit,
             vehicle_id, vehicle_number,
             dispatch_datetime, expected_delivery_datetime, actual_delivery_datetime,
             distance_km, estimated_transit_hours, order_status, transport_cost_inr)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [
        (o['Order_ID'],
         o.get('Customer_Name'), o.get('Customer_Company'), o.get('Customer_Contact'),
         o.get('Pickup_Address'), o.get('Delivery_Address'),
         o.get('Source_City'), o.get('Source_State'), o.get('Source_Pincode'),
         o.get('Destination_City'), o.get('Destination_State'), o.get('Destination_Pincode'),
         o.get('Goods_Type'), o.get('Goods_Category'), o.get('Quantity'), o.get('Unit'),
         o.get('Vehicle_ID'), o.get('Vehicle_Number'),
         o.get('Dispatch_DateTime'), o.get('Expected_Delivery_DateTime'), o.get('Actual_Delivery_DateTime'),
         o.get('Distance_km'), o.get('Estimated_Transit_Hours'),
         o.get('Order_Status'), o.get('Transport_Cost_INR'))
        for o in orders if o.get('Order_ID')
    ])
    print(f"   [OK] Orders inserted     : {len(orders)}")

    # ── Initial vehicle positions ─────────────────────────
    pos_rows = [
        (v['id'], v['current_lat'], v['current_lng'],
         v['current_speed'], v['status'], v['current_order_id'], now)
        for v in vehicles.values()
        if v['current_lat'] and v['current_lng']
    ]
    if pos_rows:
        conn.executemany("""
            INSERT INTO vehicle_positions
                (vehicle_id, lat, lng, speed, status, order_id, recorded_at)
            VALUES (?,?,?,?,?,?,?)
        """, pos_rows)
    print(f"   [OK] Positions inserted  : {len(pos_rows)}")

    # ── Alerts ────────────────────────────────────────────
    alert_rows = [
        (a['Order_ID'], order_vmap.get(a['Order_ID']),
         a.get('Alert_Type'), a.get('Alert_Reason'),
         a.get('Severity', 'Low'), a.get('Delay_Minutes', 0))
        for a in data['alerts'] if a.get('Order_ID')
    ]
    if alert_rows:
        conn.executemany("""
            INSERT INTO alerts
                (order_id, vehicle_id, alert_type, alert_reason, severity, delay_minutes)
            VALUES (?,?,?,?,?,?)
        """, alert_rows)
    print(f"   [OK] Alerts inserted     : {len(alert_rows)}")

    # ── Temperature logs ──────────────────────────────────
    temp_rows = [
        (t['Order_ID'],
         t.get('Temperature_At_Dispatch_C'), t.get('Temperature_During_Transit_C'),
         t.get('Temperature_At_Delivery_C'), t.get('Required_Temperature_Range_C'),
         t.get('Condition_Status'))
        for t in data['temperature'] if t.get('Order_ID')
    ]
    if temp_rows:
        conn.executemany("""
            INSERT INTO temperature_logs
                (order_id, temp_at_dispatch, temp_during_transit,
                 temp_at_delivery, required_range, condition_status)
            VALUES (?,?,?,?,?,?)
        """, temp_rows)
    print(f"   [OK] Temperature logs    : {len(temp_rows)}")

    # ── Documents ─────────────────────────────────────────
    doc_rows = [
        (d['Order_ID'], d.get('Invoice_Status'), d.get('Eway_Bill_Status'),
         d.get('Other_Documents'), d.get('Document_Status'))
        for d in data['documents'] if d.get('Order_ID')
    ]
    if doc_rows:
        conn.executemany("""
            INSERT INTO order_documents
                (order_id, invoice_status, eway_bill_status, other_documents, document_status)
            VALUES (?,?,?,?,?)
        """, doc_rows)
    print(f"   [OK] Documents inserted  : {len(doc_rows)}")

    conn.commit()
    conn.close()

    log_event('system', f'SQLite DB seeded from {os.path.basename(xlsx_path)}', 'system', 'seed')

    in_transit = sum(1 for o in orders if o.get('Order_Status') == 'In Transit')
    delivered  = sum(1 for o in orders if o.get('Order_Status') == 'Delivered')
    moving_v   = sum(1 for v in vehicles.values() if v['status'] == 'moving')

    print()
    print("=" * 45)
    print(f"   Orders   : {len(orders)} total  ({in_transit} In Transit, {delivered} Delivered)")
    print(f"   Vehicles : {len(vehicles)} trucks  ({moving_v} moving on map)")
    print(f"   Database : SQLite → fleet.db")
    print("=" * 45)
    print("[OK] Seeding complete!\n")

if __name__ == '__main__':
    seed()
