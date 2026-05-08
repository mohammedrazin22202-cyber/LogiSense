#!/usr/bin/env python3
"""
TrackAssist - Flask Backend with API
Intelligent Logistics Assistant
"""

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import os
import json

app = Flask(__name__)
CORS(app)

class TrackAssistBackend:
    def __init__(self, responses_dir, keywords_excel):
        self.responses_dir = responses_dir
        self.keywords_mapping = {}
        self.load_keywords_from_excel(keywords_excel)
        
    def load_keywords_from_excel(self, excel_file):
        """Load keywords from Excel file."""
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Skip header rows, start from row 3
            for row_num in range(3, ws.max_row + 1):
                filename = ws.cell(row=row_num, column=1).value
                
                if filename:
                    filename = filename.strip()
                    
                    # Get all keywords from columns 2-11
                    for col_num in range(2, 12):
                        keyword = ws.cell(row=row_num, column=col_num).value
                        if keyword:
                            keyword = keyword.strip().lower()
                            if keyword:
                                if keyword not in self.keywords_mapping:
                                    self.keywords_mapping[keyword] = []
                                if filename not in self.keywords_mapping[keyword]:
                                    self.keywords_mapping[keyword].append(filename)
            
            print(f"✓ Loaded {len(self.keywords_mapping)} keywords from Excel")
            print(f"✓ Mapped to {len(set([f for files in self.keywords_mapping.values() for f in files]))} response files")
            
        except Exception as e:
            print(f"Error loading Excel: {e}")
            raise
    
    def find_best_response(self, user_input):
        """Find best matching response for user input."""
        user_input_lower = user_input.lower()
        user_words = set(user_input_lower.split())
        
        # Score each file
        file_scores = {}
        matched_keywords_per_file = {}
        
        for keyword, files in self.keywords_mapping.items():
            keyword_words = set(keyword.split())
            
            # Exact substring match (highest priority)
            if keyword in user_input_lower:
                match_score = len(keyword_words) ** 2 * 10
                
                for file in files:
                    if file not in file_scores:
                        file_scores[file] = 0
                        matched_keywords_per_file[file] = []
                    
                    file_scores[file] += match_score
                    if keyword not in matched_keywords_per_file[file]:
                        matched_keywords_per_file[file].append(keyword)
            
            # Word overlap (secondary match)
            else:
                overlap = keyword_words & user_words
                if overlap:
                    overlap_ratio = len(overlap) / len(keyword_words)
                    
                    if overlap_ratio >= 0.5:
                        match_score = len(overlap) * overlap_ratio * 5
                        
                        for file in files:
                            if file not in file_scores:
                                file_scores[file] = 0
                                matched_keywords_per_file[file] = []
                            
                            file_scores[file] += match_score
                            if keyword not in matched_keywords_per_file[file]:
                                matched_keywords_per_file[file].append(keyword)
        
        if not file_scores:
            return None, [], 0.0
        
        # Get best match
        best_file = max(file_scores, key=file_scores.get)
        best_score = file_scores[best_file]
        matched_keywords = matched_keywords_per_file[best_file]
        
        # Calculate confidence
        max_possible_score = len(user_words) * 10
        confidence = min(best_score / max_possible_score, 1.0)
        
        # Load response
        response_path = os.path.join(self.responses_dir, best_file)
        
        if os.path.exists(response_path):
            with open(response_path, 'r', encoding='utf-8') as f:
                response = f.read().strip()
            return response, matched_keywords, confidence
        
        return None, [], 0.0

# Initialize backend
RESPONSES_DIR = './Headache'
KEYWORDS_FILE = './kword.xlsx'

# Try alternate paths

try:
    chatbot = TrackAssistBackend(RESPONSES_DIR, KEYWORDS_FILE)
    print("✓ TrackAssist backend ready!")
except Exception as e:
    print(f"❌ Error initializing backend: {e}")
    chatbot = None

@app.route('/')
def index():
    """Serve the main HTML page."""
    return send_from_directory('.', 'index.html')

@app.route('/api/chat', methods=['POST'])
def chat():
    """Handle chat messages."""
    if not chatbot:
        return jsonify({
            'error': 'Backend not initialized',
            'response': 'System error. Please contact support.'
        }), 500
    
    data = request.get_json()
    user_message = data.get('message', '').strip()
    
    if not user_message:
        return jsonify({
            'response': "Hello! I'm TrackAssist. I can help you with order tracking, shipment information, and questions about our services."
        })
    
    response, keywords, confidence = chatbot.find_best_response(user_message)
    
    if response and confidence > 0.3:
        return jsonify({
            'response': response,
            'keywords': keywords[:3],
            'confidence': confidence
        })
    else:
        return jsonify({
            'response': ("I apologize, but I need more context to assist you properly. Could you rephrase your question?\n\n"
                        "I can help with:\n"
                        "📦 Order tracking and shipment status\n"
                        "🆕 Placing new orders and bookings\n"
                        "🏢 Company information and services\n"
                        "🐟 Seafood and cold chain logistics\n"
                        "🚛 Fleet, trucks, and vehicle information\n"
                        "✅ Compliance, certifications, and safety\n"
                        "🔧 System issues and tracking problems\n\n"
                        "Feel free to ask about any of these topics!")
        })

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'keywords_loaded': len(chatbot.keywords_mapping) if chatbot else 0,
        'backend': 'ready' if chatbot else 'error'
    })

if __name__ == '__main__':
    print("\n" + "="*70)
    print(" "*22 + "TrackAssist Backend")
    print(" "*15 + "Intelligent Logistics Assistant")
    print("="*70)
    print("\nStarting Flask server...")
    print("Server will be available at: http://localhost:5000")
    print("Press Ctrl+C to stop\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
